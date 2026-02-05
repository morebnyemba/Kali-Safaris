# exports.py

import openpyxl
from openpyxl.styles import Font
from openpyxl.cell import MergedCell
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from decimal import Decimal
from django.conf import settings
import logging

# For PDF generation
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from .models import Payment, Booking, Traveler

logger = logging.getLogger(__name__)

# --- Helper Functions ---

def _get_church_name():
    """Gets the company name from settings, with a fallback."""
    try:
        # Use the new centralized COMPANY_DETAILS setting
        return settings.COMPANY_DETAILS['NAME']
    except (AttributeError, KeyError):
        logger.warning("settings.COMPANY_DETAILS['NAME'] not found. Using fallback 'Company'.")
        return "Company"

def _auto_adjust_excel_columns(sheet):
    """
    Auto-adjusts the width of columns in an Excel sheet based on content length.
    This version correctly handles merged cells by iterating over rows and skipping them.
    """
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue  # Ignore merged cells

            if cell.value:
                # Get current max width for this column and update if needed
                current_max = column_widths.get(cell.column_letter, 0)
                column_widths[cell.column_letter] = max(current_max, len(str(cell.value)))

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width + 2  # Add padding

def _draw_pdf_footer(canvas, doc):
    """Draws a standard footer on each PDF page with company details."""
    canvas.saveState()
    
    # Get details from settings, with fallbacks
    details = settings.COMPANY_DETAILS
    name = details.get('NAME', 'Our Company')
    address_line_1 = details.get('ADDRESS_LINE_1', '')
    address_line_2 = details.get('ADDRESS_LINE_2', '')
    phone = details.get('CONTACT_PHONE', '')
    email = details.get('CONTACT_EMAIL', '')
    website = details.get('WEBSITE', '')

    # Combine address parts that exist
    address_parts = [part for part in [address_line_1, address_line_2] if part]
    full_address = ", ".join(address_parts)

    # Combine contact parts that exist
    contact_parts = []
    if phone: contact_parts.append(f"Phone: {phone}")
    if email: contact_parts.append(f"Email: {email}")
    if website: contact_parts.append(f"Website: {website}")
    contact_line = " | ".join(contact_parts)

    canvas.setFont('Helvetica', 8)
    
    line1_text = f"{name} - {full_address}" if full_address else name
    y1 = doc.bottomMargin - 12
    canvas.drawCentredString(doc.width / 2 + doc.leftMargin, y1, line1_text)

    canvas.setFont('Helvetica', 8)
    y2 = doc.bottomMargin - 24
    canvas.drawCentredString(doc.width / 2 + doc.leftMargin, y2, contact_line)

    canvas.setFont('Helvetica-Oblique', 7)
    y3 = doc.bottomMargin - 34
    canvas.drawCentredString(doc.width / 2 + doc.leftMargin, y3, "Powered by Slyker Tech Web Services")

    canvas.restoreState()

def get_giver_name(payment):
    """Helper to get the best available name for a giver."""
    # Payment is linked to a booking, which has a customer
    if payment.booking and payment.booking.customer:
        customer = payment.booking.customer
        full_name = customer.get_full_name()
        if full_name:
            return full_name
        elif customer.contact and customer.contact.name:
            return customer.contact.name
        elif customer.contact:
            return customer.contact.whatsapp_id
    return "Anonymous Giver"

# --- Payment Summary Export Functions ---

def export_payment_summary_to_excel(queryset, period_name):
    """
    Generates an Excel file summarizing payments by type for a given period.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="payment_summary_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Payment Summary - {period_name.title()}'

    church_name = _get_church_name()
    main_title_font = Font(bold=True, size=16)
    sheet.cell(row=1, column=1, value=church_name).font = main_title_font
    sheet.merge_cells('A1:C1')

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=2, column=1, value=f"Payment Summary for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A2:C2')

    powered_by_font = Font(italic=True, size=9)
    sheet.cell(row=3, column=1, value="Powered by Slyker Tech Web Services").font = powered_by_font
    sheet.merge_cells('A3:C3')

    headers = ["Payment Type", "Total Amount (USD)", "Transaction Count"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=5, column=col_num, value=header)
        cell.font = header_font

    summary_data = queryset.values('payment_type').annotate(total_amount=Sum('amount'), transaction_count=Count('id')).order_by('payment_type')

    row_num = 6
    grand_total_amount = Decimal('0.00')
    grand_total_count = 0
    payment_type_display_map = dict(Payment.PAYMENT_TYPE_CHOICES)

    for summary in summary_data:
        payment_type_key = summary['payment_type']
        # Explicitly convert lazy translation object to a string for openpyxl
        payment_type_display = str(payment_type_display_map.get(payment_type_key, payment_type_key.title()))
        sheet.cell(row=row_num, column=1, value=payment_type_display)

        cell_amount = sheet.cell(row=row_num, column=2, value=summary['total_amount'])
        cell_amount.number_format = '"$"#,##0.00'

        sheet.cell(row=row_num, column=3, value=summary['transaction_count'])
        grand_total_amount += summary['total_amount']
        grand_total_count += summary['transaction_count']
        row_num += 1

    sheet.cell(row=row_num, column=1, value="Grand Total").font = header_font
    cell_grand_total = sheet.cell(row=row_num, column=2, value=grand_total_amount)
    cell_grand_total.font = header_font
    cell_grand_total.number_format = '"$"#,##0.00'
    sheet.cell(row=row_num, column=3, value=grand_total_count).font = header_font

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_payment_summary_to_pdf(queryset, period_name):
    """
    Generates a PDF file summarizing payments by type for a given period.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    church_name = _get_church_name()
    title = Paragraph(f"{church_name} - Payment Summary: {period_name.replace('_', ' ').title()}", styles['h1'])
    subtitle = Paragraph(f"Report generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
    elements.extend([title, subtitle, Spacer(1, 24)])

    summary_data = queryset.values('payment_type').annotate(total_amount=Sum('amount'), transaction_count=Count('id')).order_by('payment_type')

    headers = ["Payment Type", "Total Amount (USD)", "Transactions"]
    data = [headers]
    grand_total_amount = Decimal('0.00')
    grand_total_count = 0
    payment_type_display_map = dict(Payment.PAYMENT_TYPE_CHOICES)

    for summary in summary_data:
        payment_type_key = summary['payment_type']
        total_amount = summary['total_amount']
        # Explicitly convert lazy translation object to a string for safety
        payment_type_display = str(payment_type_display_map.get(payment_type_key, payment_type_key.title()))
        data.append([payment_type_display, f"${total_amount:,.2f}", str(summary['transaction_count'])])
        grand_total_amount += total_amount
        grand_total_count += summary['transaction_count']

    data.append([Paragraph("<b>Grand Total</b>", styles['Normal']), Paragraph(f"<b>${grand_total_amount:,.2f}</b>", styles['Normal']), Paragraph(f"<b>{grand_total_count}</b>", styles['Normal'])])

    table = Table(data, colWidths=[200, 150, 100], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C0C0C0')),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payment_summary_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response

# --- Givers List Export Functions ---

def export_givers_list_finance_excel(queryset, period_name):
    """
    Generates an Excel file listing all givers and their total contributions for a period.
    For the finance department.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="givers_finance_report_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Givers Report (Finance)'

    church_name = _get_church_name()
    main_title_font = Font(bold=True, size=16)
    sheet.cell(row=1, column=1, value=church_name).font = main_title_font
    sheet.merge_cells('A1:B1')

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=2, column=1, value=f"Givers Report (Finance) for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A2:B2')

    powered_by_font = Font(italic=True, size=9)
    sheet.cell(row=3, column=1, value="Powered by Slyker Tech Web Services").font = powered_by_font
    sheet.merge_cells('A3:B3')

    headers = ["Giver Name", "Total Amount (USD)"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=5, column=col_num, value=header)
        cell.font = header_font

    givers = {}
    for payment in queryset.select_related('booking__customer__contact'):
        # Get customer from booking
        if not payment.booking or not payment.booking.customer:
            continue
        customer_id = payment.booking.customer.contact_id
        if customer_id not in givers:
            givers[customer_id] = {'name': get_giver_name(payment), 'total': Decimal('0.00')}
        givers[customer_id]['total'] += payment.amount

    sorted_givers = sorted(givers.values(), key=lambda x: x['name'])
    
    row_num = 6
    for giver in sorted_givers:
        sheet.cell(row=row_num, column=1, value=giver['name'])
        sheet.cell(row=row_num, column=2, value=giver['total']).number_format = '"$"#,##0.00'
        row_num += 1

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_givers_list_finance_pdf(queryset, period_name):
    """
    Generates a PDF file listing all givers and their total contributions for a period.
    For the finance department.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    church_name = _get_church_name()
    title = Paragraph(f"{church_name} - Givers Report (Finance): {period_name.replace('_', ' ').title()}", styles['h1'])
    elements.extend([title, Spacer(1, 24)])

    givers = {}
    for payment in queryset.select_related('booking__customer__contact'):
        # Get customer from booking
        if not payment.booking or not payment.booking.customer:
            continue
        customer_id = payment.booking.customer.contact_id
        if customer_id not in givers:
            givers[customer_id] = {'name': get_giver_name(payment), 'total': Decimal('0.00')}
        givers[customer_id]['total'] += payment.amount

    sorted_givers = sorted(givers.values(), key=lambda x: x['name'])

    headers = ["Giver Name", "Total Amount (USD)"]
    data = [headers]
    for giver in sorted_givers:
        data.append([giver['name'], f"${giver['total']:,.2f}"])

    table = Table(data, colWidths=[300, 150], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="givers_finance_report_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response

def export_givers_list_publication_excel(queryset, period_name):
    """
    Generates an Excel file listing the names of all givers for public acknowledgment.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="givers_publication_list_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Givers List (Publication)'

    church_name = _get_church_name()
    main_title_font = Font(bold=True, size=16)
    sheet.cell(row=1, column=1, value=church_name).font = main_title_font
    sheet.merge_cells('A1:A1')

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=2, column=1, value=f"Thank You To Our Givers for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A2:A2')

    powered_by_font = Font(italic=True, size=9)
    sheet.cell(row=3, column=1, value="Powered by Slyker Tech Web Services").font = powered_by_font
    sheet.merge_cells('A3:A3')

    sheet.cell(row=5, column=1, value="Giver Name").font = header_font

    giver_names = sorted(list(set(get_giver_name(p) for p in queryset.select_related('booking__customer__contact'))))

    for row_num, name in enumerate(giver_names, 6):
        sheet.cell(row=row_num, column=1, value=name)

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_givers_list_publication_pdf(queryset, period_name):
    """
    Generates a PDF file listing the names of all givers for public acknowledgment.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    church_name = _get_church_name()
    title = Paragraph(f"{church_name} - A Special Thank You To Our Givers", styles['h1'])
    subtitle = Paragraph(f"For {period_name.replace('_', ' ').title()}", styles['h2'])
    elements.extend([title, subtitle, Spacer(1, 24)])

    giver_names = sorted(list(set(get_giver_name(p) for p in queryset.select_related('booking__customer__contact'))))

    data = [["Giver Name"]] + [[name] for name in giver_names]

    table = Table(data, colWidths=[450], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="givers_publication_list_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response


# --- Booking Manifest Export Functions ---

def export_passenger_manifest_summary_excel(booking_date):
    """
    Generates an operational summary report in Excel format for confirmed bookings on a specific date.
    Shows headcount per booking group for crew seating arrangements and park fees.
    
    Args:
        booking_date: Date object for which to generate the summary
        
    Returns:
        HttpResponse with Excel content
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"passenger_summary_{booking_date.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Passenger Summary'
    
    # Company name and title
    company_name = getattr(settings, 'COMPANY_DETAILS', {}).get('NAME', 'Kalai Safaris')
    main_title_font = Font(bold=True, size=16)
    sheet.cell(row=1, column=1, value=company_name).font = main_title_font
    sheet.merge_cells('A1:E1')
    
    bold_font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value="Passenger Manifest Summary - Operational Report").font = bold_font
    sheet.merge_cells('A2:E2')
    
    # Date info
    sheet.cell(row=3, column=1, value=f"Tour Date: {booking_date.strftime('%B %d, %Y')}")
    sheet.merge_cells('A3:E3')
    
    sheet.cell(row=4, column=1, value=f"Generated: {timezone.now().strftime('%B %d, %Y at %H:%M')}")
    sheet.merge_cells('A4:E4')
    
    powered_by_font = Font(italic=True, size=9)
    sheet.cell(row=5, column=1, value="Powered by Slyker Tech Web Services").font = powered_by_font
    sheet.merge_cells('A5:E5')
    
    # Get confirmed bookings
    confirmed_bookings = Booking.objects.filter(
        start_date=booking_date,
        payment_status__in=[Booking.PaymentStatus.PAID, Booking.PaymentStatus.DEPOSIT_PAID]
    ).prefetch_related('travelers').select_related('customer').order_by('booking_reference')
    
    if not confirmed_bookings.exists():
        sheet.cell(row=7, column=1, value=f"No confirmed bookings found for {booking_date.strftime('%B %d, %Y')}")
    else:
        # Build summary data
        booking_groups = []
        total_passengers = 0
        total_with_id_docs = 0
        
        for booking in confirmed_bookings:
            travelers = booking.travelers.all()
            traveler_count = travelers.count()
            if traveler_count == 0:
                traveler_count = booking.number_of_adults + booking.number_of_children
            
            # Count travelers with ID documents
            id_docs_count = sum(1 for t in travelers if t.id_document)
            
            if booking.customer:
                customer_name = booking.customer.get_full_name() or (booking.customer.contact.name if booking.customer.contact else booking.booking_reference)
            else:
                customer_name = booking.booking_reference
            
            booking_groups.append({
                'name': customer_name,
                'reference': booking.booking_reference,
                'count': traveler_count,
                'tour': booking.tour_name,
                'id_docs': id_docs_count
            })
            total_passengers += traveler_count
            total_with_id_docs += id_docs_count
        
        # Summary in requested format
        summary_parts = [f"{group['name']}({group['count']})" for group in booking_groups]
        summary_text = ", ".join(summary_parts) + f" total: {total_passengers}"
        
        sheet.cell(row=7, column=1, value="Confirmed Passenger Summary:").font = Font(bold=True)
        sheet.merge_cells('A7:D7')
        
        sheet.cell(row=8, column=1, value=summary_text)
        sheet.merge_cells('A8:E8')
        
        sheet.cell(row=9, column=1, value=f"ID Documents Uploaded: {total_with_id_docs} of {total_passengers}")
        sheet.merge_cells('A9:E9')
        
        # Breakdown table
        header_font = Font(bold=True)
        headers = ["Booking Reference", "Customer/Group", "Tour", "Passengers", "ID Docs"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=11, column=col_num, value=header)
            cell.font = header_font
        
        row_num = 12
        for group in booking_groups:
            sheet.cell(row=row_num, column=1, value=group['reference'])
            sheet.cell(row=row_num, column=2, value=group['name'])
            sheet.cell(row=row_num, column=3, value=group['tour'])
            sheet.cell(row=row_num, column=4, value=group['count'])
            sheet.cell(row=row_num, column=5, value=f"{group['id_docs']}/{group['count']}")
            row_num += 1
        
        # Total row
        sheet.cell(row=row_num, column=1, value="TOTAL").font = header_font
        sheet.cell(row=row_num, column=4, value=total_passengers).font = header_font
        sheet.cell(row=row_num, column=5, value=f"{total_with_id_docs}/{total_passengers}").font = header_font
        
        # Usage notes
        row_num += 2
        sheet.cell(row=row_num, column=1, value="Usage Notes:").font = Font(bold=True)
        sheet.merge_cells(f'A{row_num}:E{row_num}')
        row_num += 1
        sheet.cell(row=row_num, column=1, value="• Individual counts - Used by crew for seating arrangements and logistics")
        sheet.merge_cells(f'A{row_num}:E{row_num}')
        row_num += 1
        sheet.cell(row=row_num, column=1, value="• Total count - Used for calculating park fees and vehicle capacity")
        sheet.merge_cells(f'A{row_num}:E{row_num}')
        row_num += 1
        sheet.cell(row=row_num, column=1, value="• ID Docs - Number of travelers with ID/Passport images uploaded via WhatsApp")
        sheet.merge_cells(f'A{row_num}:E{row_num}')
    
    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response


def export_passenger_manifest_summary_pdf(booking_date):
    """
    Generates an operational summary report for confirmed bookings on a specific date.
    Shows headcount per booking group for crew seating arrangements and park fees.
    
    Format: Customer/Group(count), Customer/Group(count), total: X
    
    Args:
        booking_date: Date object for which to generate the summary
        
    Returns:
        HttpResponse with PDF content
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=50
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Get company details from settings
    company_name = getattr(settings, 'COMPANY_DETAILS', {}).get('NAME', 'Kalai Safaris')
    
    # Title
    title = Paragraph(
        f"<b>{company_name}</b><br/>Passenger Manifest Summary - Operational Report",
        styles['h1']
    )
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date information
    date_info = Paragraph(
        f"<b>Tour Date:</b> {booking_date.strftime('%B %d, %Y')}<br/>"
        f"<b>Report Generated:</b> {timezone.now().strftime('%B %d, %Y at %H:%M')}",
        styles['Normal']
    )
    elements.append(date_info)
    elements.append(Spacer(1, 20))
    
    # Get only confirmed (paid/deposit paid) bookings for the specified date
    confirmed_bookings = Booking.objects.filter(
        start_date=booking_date,
        payment_status__in=[Booking.PaymentStatus.PAID, Booking.PaymentStatus.DEPOSIT_PAID]
    ).prefetch_related('travelers').select_related('customer').order_by('booking_reference')
    
    if not confirmed_bookings.exists():
        no_bookings_text = Paragraph(
            f"<i>No confirmed bookings found for {booking_date.strftime('%B %d, %Y')}</i>",
            styles['Italic']
        )
        elements.append(no_bookings_text)
    else:
        # Build summary data
        booking_groups = []
        total_passengers = 0
        
        for booking in confirmed_bookings:
            traveler_count = booking.travelers.count()
            # If no travelers recorded, use the booking's adult + children count
            if traveler_count == 0:
                traveler_count = booking.number_of_adults + booking.number_of_children
            
            # Get customer name or use booking reference
            if booking.customer:
                customer_name = booking.customer.get_full_name() or booking.customer.contact.name if booking.customer.contact else booking.booking_reference
            else:
                customer_name = booking.booking_reference
            
            booking_groups.append({
                'name': customer_name,
                'reference': booking.booking_reference,
                'count': traveler_count,
                'tour': booking.tour_name
            })
            total_passengers += traveler_count
        
        # Summary paragraph in requested format
        summary_parts = [f"{group['name']}({group['count']})" for group in booking_groups]
        summary_text = ", ".join(summary_parts) + f" <b>total: {total_passengers}</b>"
        
        summary_para = Paragraph(
            f"<b>Confirmed Passenger Summary:</b><br/>{summary_text}<br/>"
            f"<b>ID Documents Uploaded:</b> {total_with_id_docs} of {total_passengers}",
            styles['Normal']
        )
        elements.append(summary_para)
        elements.append(Spacer(1, 20))
        
        # Detailed breakdown table
        breakdown_title = Paragraph("<b>Breakdown by Booking Group:</b>", styles['Heading2'])
        elements.append(breakdown_title)
        elements.append(Spacer(1, 10))
        
        # Build table data
        headers = ["Booking Reference", "Customer/Group", "Tour", "Passengers", "ID Docs"]
        data = [headers]
        
        for group in booking_groups:
            data.append([
                group['reference'],
                group['name'],
                group['tour'],
                str(group['count']),
                f"{group['id_docs']}/{group['count']}"
            ])
        
        # Add total row
        data.append([
            Paragraph("<b>TOTAL</b>", styles['Normal']),
            "",
            "",
            Paragraph(f"<b>{total_passengers}</b>", styles['Normal']),
            Paragraph(f"<b>{total_with_id_docs}/{total_passengers}</b>", styles['Normal'])
        ])
        
        # Create table
        col_widths = [120, 130, 130, 70, 70]
        table = Table(data, colWidths=col_widths, hAlign='LEFT')
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F2D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
            ('ALIGN', (3, 1), (-1, -2), 'CENTER'),
            
            # Total row styling
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D3D3D3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (3, -1), (3, -1), 'CENTER'),
            
            # Grid and padding
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)
        
        # Add usage notes
        elements.append(Spacer(1, 20))
        notes = Paragraph(
            "<b>Usage Notes:</b><br/>"
            "• <i>Individual counts</i> - Used by crew for seating arrangements and logistics<br/>"
            "• <i>Total count</i> - Used for calculating park fees and vehicle capacity<br/>"
            "• <i>ID Docs</i> - Number of travelers with ID/Passport images uploaded via WhatsApp",
            styles['Normal']
        )
        elements.append(notes)
    
    # Build PDF
    doc.build(elements, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"passenger_summary_{booking_date.strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_booking_manifest_pdf(booking_date):
    """
    Generates a professional PDF manifest for bookings on a specific date.
    Contains only essential information needed by ZimParks:
    - Full Name
    - ID Number
    - Nationality
    - Age
    
    Args:
        booking_date: Date object for which to generate the manifest
        
    Returns:
        HttpResponse with PDF content
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=40, 
        bottomMargin=50
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Get company details from settings
    company_name = getattr(settings, 'COMPANY_DETAILS', {}).get('NAME', 'Kalai Safaris')
    
    # Title
    title = Paragraph(
        f"<b>{company_name}</b><br/>Safari Booking Manifest", 
        styles['h1']
    )
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date information
    date_info = Paragraph(
        f"<b>Tour Date:</b> {booking_date.strftime('%B %d, %Y')}<br/>"
        f"<b>Report Generated:</b> {timezone.now().strftime('%B %d, %Y at %H:%M')}",
        styles['Normal']
    )
    elements.append(date_info)
    elements.append(Spacer(1, 20))
    
    # Get all bookings for the specified date
    bookings = Booking.objects.filter(
        start_date=booking_date
    ).prefetch_related('travelers').order_by('booking_reference')
    
    if not bookings.exists():
        no_bookings_text = Paragraph(
            f"<i>No bookings found for {booking_date.strftime('%B %d, %Y')}</i>",
            styles['Italic']
        )
        elements.append(no_bookings_text)
    else:
        # Summary
        total_travelers = sum(booking.travelers.count() for booking in bookings)
        summary = Paragraph(
            f"<b>Total Bookings:</b> {bookings.count()}<br/>"
            f"<b>Total Travelers:</b> {total_travelers}",
            styles['Normal']
        )
        elements.append(summary)
        elements.append(Spacer(1, 20))
        
        # Build table data
        headers = ["Full Name", "ID Number", "Nationality", "Age", "ID Doc"]
        data = [headers]
        
        for booking in bookings:
            # Add booking reference header
            booking_header = [
                Paragraph(f"<b>Booking: {booking.booking_reference} - {booking.tour_name}</b>", styles['Normal']),
                "", "", "", ""
            ]
            data.append(booking_header)
            
            # Add travelers for this booking
            travelers = booking.travelers.all().order_by('traveler_type', 'name')
            if not travelers.exists():
                # If no traveler details, show booking info
                data.append([
                    "No traveler details recorded",
                    "-",
                    "-",
                    "-",
                    "-"
                ])
            else:
                for traveler in travelers:
                    # Check if ID document is uploaded
                    id_doc_status = "✓" if traveler.id_document else "✗"
                    data.append([
                        traveler.name,
                        traveler.id_number or "Not provided",
                        traveler.nationality or "Not provided",
                        str(traveler.age),
                        id_doc_status
                    ])
            
            # Add spacing row between bookings
            data.append(["", "", "", "", ""])
        
        # Create table with appropriate column widths
        col_widths = [180, 110, 90, 50, 50]
        table = Table(data, colWidths=col_widths, hAlign='LEFT')
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F2D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid lines
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors for better readability
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)
        
        # Add note about ID document column
        elements.append(Spacer(1, 10))
        id_note = Paragraph(
            "<b>ID Doc:</b> ✓ = ID/Passport image uploaded via WhatsApp | ✗ = Not uploaded",
            styles['Normal']
        )
        elements.append(id_note)
    
    # Build PDF
    doc.build(elements, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"booking_manifest_{booking_date.strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
